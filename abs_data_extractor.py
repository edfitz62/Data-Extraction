# -*- coding: utf-8 -*-
"""
Created on Fri Jul 11 11:52:51 2025

@author: edfit
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from typing import Dict, List, Tuple, Any
import logging
import re

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ABSDataExtractor:
    """
    A comprehensive class to extract ABS performance data from surveillance dashboards
    and organize it into normalized dataframes for database storage.
    """
    
    def __init__(self):
        self.deal_master_data = []
        self.collateral_characteristics_data = []
        self.performance_metrics_data = []
        self.historical_performance_data = []
        self.cnl_analysis_data = []
        self.trigger_analysis_data = []
        self.statistical_summaries_data = []
        
        # Define metric mappings for different sections
        self.metric_mappings = {
            'deal_master': [
                'Deal Short Name', 'KBRA Subsector', 'Securitization Date',
                'Prior Action Date', 'Distribution Date', 'Months Seasoned'
            ],
            'collateral_snapshot': [
                'Original Collat Bal ($\'000s)', 'Current Collat Bal ($\'000s)',
                'Pool Factor', 'WA Interest Rate (At Closing)', 'WA Interest Rate (Current)',
                'WA Remaining Term\r\n (At Closing)', 'WA Remaining Term (Current)',
                'Original # of Receivables ', 'Current Number of Receivables'
            ],
            'current_performance': [
                '30+ DQ ', '60+ DQ', '90+ DQ', 'Extensions', 'CNL'
            ],
            'cnl_analysis': [
                'CNL', 'Expected Base Case at Closing (for current period)',
                'Base Case Loss Projection (at Closing)', 'Base Case Loss Projection (Current)'
            ],
            'trigger_analysis': [
                'CNL Trigger Level', 'Current Trigger Breach', 'Trigger Cushion'
            ],
            'overcollateralization': [
                'Intial ', 'T0 (Current Distribution Date)', 'T-3', 'T-6', 'T-9', 'T-12', 'Target'
            ],
            'reserve_account': [
                'Intial', 'Current (% of Original Deal Bal)', 'Current (% of Current Deal Bal)', 'Target '
            ]
        }
    
    def extract_from_files(self, octane_file: str, carvana_file: str) -> Dict[str, pd.DataFrame]:
        """
        Main extraction method that processes both files and returns all dataframes.
        """
        logger.info("Starting ABS data extraction process")
        
        # Process Octane file
        logger.info("Processing Octane file...")
        self._process_file(octane_file, "Octane")
        
        # Process Carvana file
        logger.info("Processing Carvana file...")
        self._process_file(carvana_file, "Carvana")
        
        # Convert collected data to DataFrames
        dataframes = self._create_dataframes()
        
        logger.info("Data extraction completed successfully")
        return dataframes
    
    def _process_file(self, file_path: str, issuer_name: str):
        """
        Process a single Excel file and extract data from Deal Level - Vertical sheet.
        """
        try:
            # Read the Excel file
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract from Deal Level - Vertical sheet
            if "Deal Level - Vertical" in workbook.sheetnames:
                self._extract_from_vertical_sheet(workbook["Deal Level - Vertical"], issuer_name)
            else:
                logger.warning(f"Deal Level - Vertical sheet not found in {file_path}")
                
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {str(e)}")
            raise
    
    def _extract_from_vertical_sheet(self, worksheet, issuer_name: str):
        """
        Extract data from the Deal Level - Vertical worksheet.
        """
        # Read all data from worksheet
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))
        
        # Find deal names in row 2 (index 1)
        deal_names = []
        header_row = data[1] if len(data) > 1 else []
        
        for i, cell in enumerate(header_row[2:], start=2):  # Skip first two columns
            if cell and isinstance(cell, str) and cell.strip():
                deal_names.append((cell.strip(), i))
        
        if not deal_names:
            logger.warning(f"No deal names found in {issuer_name} file")
            return
        
        logger.info(f"Found {len(deal_names)} deals in {issuer_name}: {[name for name, _ in deal_names]}")
        
        # Extract data for each deal
        for deal_name, col_idx in deal_names:
            self._extract_deal_data(data, deal_name, col_idx, issuer_name)
    
    def _extract_deal_data(self, data: List[List], deal_name: str, col_idx: int, issuer_name: str):
        """
        Extract all metrics for a single deal.
        """
        current_section = ""
        as_of_date = datetime.now().date()  # Default to current date
        
        # Initialize deal record
        deal_record = {
            'issuer_name': issuer_name,
            'deal_short_name': deal_name,
            'as_of_date': as_of_date
        }
        
        # Parse through all rows to extract metrics
        for row_idx, row in enumerate(data):
            if len(row) <= col_idx:
                continue
                
            # Check if this is a section header (has value in column 1 but not column 0)
            if len(row) > 1 and row[1] and not row[0]:
                current_section = str(row[1]).strip()
                continue
            
            # Check if this is a data row (has values in both column 0 and 1)
            if len(row) > 1 and row[0] and row[1]:
                metric_name = str(row[1]).strip()
                metric_value = row[col_idx] if col_idx < len(row) else None
                
                # Extract data based on current section
                self._process_metric(deal_record, current_section, metric_name, metric_value)
        
        # Store the extracted data in appropriate collections
        self._store_deal_data(deal_record, current_section)
    
    def _process_metric(self, deal_record: Dict, section: str, metric_name: str, metric_value: Any):
        """
        Process a single metric and add it to the appropriate deal record.
        """
        if metric_value is None:
            return
        
        # Handle different data types
        if isinstance(metric_value, datetime):
            metric_value = metric_value.date()
        elif isinstance(metric_value, (int, float)):
            metric_value = float(metric_value)
        elif isinstance(metric_value, str):
            metric_value = metric_value.strip()
        
        # Map metrics to appropriate categories
        if metric_name in self.metric_mappings.get('deal_master', []):
            self._add_deal_master_metric(deal_record, metric_name, metric_value)
        elif metric_name in self.metric_mappings.get('collateral_snapshot', []):
            self._add_collateral_metric(deal_record, metric_name, metric_value)
        elif metric_name in self.metric_mappings.get('current_performance', []):
            self._add_performance_metric(deal_record, metric_name, metric_value)
        elif section == "CNL Analysis - Actual, Base Case":
            self._add_cnl_metric(deal_record, metric_name, metric_value)
        elif section == "Trigger Analysis":
            self._add_trigger_metric(deal_record, metric_name, metric_value)
        elif "Delinquencies" in section and metric_name.startswith(('T0', 'T-', 'LTM')):
            self._add_historical_metric(deal_record, section, metric_name, metric_value)
        elif section == "Extensions" and metric_name.startswith(('T0', 'T-', 'LTM')):
            self._add_historical_metric(deal_record, section, metric_name, metric_value)
        elif section == "Overcollateralization":
            self._add_oc_metric(deal_record, metric_name, metric_value)
        elif section == "Reserve Account":
            self._add_reserve_metric(deal_record, metric_name, metric_value)
    
    def _add_deal_master_metric(self, deal_record: Dict, metric_name: str, metric_value: Any):
        """Add deal master metrics."""
        metric_map = {
            'Deal Short Name': 'deal_short_name',
            'KBRA Subsector': 'kbra_subsector',
            'Securitization Date': 'securitization_date',
            'Prior Action Date': 'prior_action_date',
            'Distribution Date': 'distribution_date',
            'Months Seasoned': 'months_seasoned'
        }
        
        if metric_name in metric_map:
            deal_record[metric_map[metric_name]] = metric_value
    
    def _add_collateral_metric(self, deal_record: Dict, metric_name: str, metric_value: Any):
        """Add collateral characteristics metrics."""
        metric_map = {
            'Original Collat Bal ($\'000s)': 'original_collateral_balance',
            'Current Collat Bal ($\'000s)': 'current_collateral_balance',
            'Pool Factor': 'pool_factor',
            'WA Interest Rate (At Closing)': 'wa_interest_rate_closing',
            'WA Interest Rate (Current)': 'wa_interest_rate_current',
            'WA Remaining Term\r\n (At Closing)': 'wa_remaining_term_closing',
            'WA Remaining Term (Current)': 'wa_remaining_term_current',
            'Original # of Receivables ': 'original_num_receivables',
            'Current Number of Receivables': 'current_num_receivables'
        }
        
        if metric_name in metric_map:
            deal_record[metric_map[metric_name]] = metric_value
    
    def _add_performance_metric(self, deal_record: Dict, metric_name: str, metric_value: Any):
        """Add current performance metrics."""
        metric_map = {
            '30+ DQ ': 'delinquency_30_plus',
            '60+ DQ': 'delinquency_60_plus',
            '90+ DQ': 'delinquency_90_plus',
            'Extensions': 'extensions_rate',
            'CNL': 'cnl_rate'
        }
        
        if metric_name in metric_map:
            deal_record[metric_map[metric_name]] = metric_value
    
    def _add_cnl_metric(self, deal_record: Dict, metric_name: str, metric_value: Any):
        """Add CNL analysis metrics."""
        metric_map = {
            'CNL': 'actual_cnl',
            'Expected Base Case at Closing (for current period)': 'expected_base_case_closing',
            'Base Case Loss Projection (at Closing)': 'base_case_projection_closing',
            'Base Case Loss Projection (Current)': 'base_case_projection_current'
        }
        
        if metric_name in metric_map:
            deal_record[f'cnl_{metric_map[metric_name]}'] = metric_value
    
    def _add_trigger_metric(self, deal_record: Dict, metric_name: str, metric_value: Any):
        """Add trigger analysis metrics."""
        metric_map = {
            'CNL Trigger Level': 'cnl_trigger_level',
            'Current Trigger Breach': 'current_trigger_breach',
            'Trigger Cushion': 'trigger_cushion'
        }
        
        if metric_name in metric_map:
            deal_record[f'trigger_{metric_map[metric_name]}'] = metric_value
    
    def _add_historical_metric(self, deal_record: Dict, section: str, metric_name: str, metric_value: Any):
        """Add historical performance metrics."""
        # Determine metric type from section
        if "30+" in section:
            metric_type = "30+DQ"
        elif "60+" in section:
            metric_type = "60+DQ"
        elif "90+" in section:
            metric_type = "90+DQ"
        elif "Extensions" in section:
            metric_type = "Extensions"
        else:
            metric_type = section
        
        # Handle different time periods
        if metric_name.startswith('T'):
            period_offset = metric_name.split(' ')[0]  # T0, T-3, etc.
            
            historical_record = {
                'issuer_name': deal_record['issuer_name'],
                'deal_short_name': deal_record['deal_short_name'],
                'as_of_date': deal_record['as_of_date'],
                'metric_type': metric_type,
                'period_offset': period_offset,
                'metric_value': metric_value
            }
            self.historical_performance_data.append(historical_record)
        
        elif metric_name.startswith('LTM'):
            # Handle LTM statistics
            stat_type = metric_name.replace('LTM ', '').lower()
            
            stats_record = {
                'issuer_name': deal_record['issuer_name'],
                'deal_short_name': deal_record['deal_short_name'],
                'as_of_date': deal_record['as_of_date'],
                'metric_type': metric_type,
                'stat_type': stat_type,
                'stat_value': metric_value
            }
            self.statistical_summaries_data.append(stats_record)
    
    def _add_oc_metric(self, deal_record: Dict, metric_name: str, metric_value: Any):
        """Add overcollateralization metrics."""
        if metric_name == 'T0 (Current Distribution Date)':
            deal_record['oc_current'] = metric_value
        elif metric_name == 'Target':
            deal_record['oc_target'] = metric_value
        elif metric_name == 'Intial ':
            deal_record['oc_initial'] = metric_value
    
    def _add_reserve_metric(self, deal_record: Dict, metric_name: str, metric_value: Any):
        """Add reserve account metrics."""
        metric_map = {
            'Intial': 'reserve_initial',
            'Current (% of Original Deal Bal)': 'reserve_pct_original',
            'Current (% of Current Deal Bal)': 'reserve_pct_current',
            'Target ': 'reserve_target'
        }
        
        if metric_name in metric_map:
            deal_record[metric_map[metric_name]] = metric_value
    
    def _store_deal_data(self, deal_record: Dict, current_section: str):
        """Store deal data in appropriate collections."""
        # Deal Master Data
        master_data = {k: v for k, v in deal_record.items() 
                      if k in ['issuer_name', 'deal_short_name', 'as_of_date', 'kbra_subsector',
                              'securitization_date', 'prior_action_date', 'distribution_date',
                              'months_seasoned', 'original_collateral_balance', 'current_collateral_balance',
                              'pool_factor', 'original_num_receivables', 'current_num_receivables']}
        
        if len(master_data) > 3:  # More than just the basic keys
            self.deal_master_data.append(master_data)
        
        # Collateral Characteristics Data
        collateral_data = {k: v for k, v in deal_record.items() 
                          if k in ['issuer_name', 'deal_short_name', 'as_of_date',
                                  'wa_interest_rate_closing', 'wa_interest_rate_current',
                                  'wa_remaining_term_closing', 'wa_remaining_term_current']}
        
        if len(collateral_data) > 3:
            self.collateral_characteristics_data.append(collateral_data)
        
        # Performance Metrics Data
        performance_data = {k: v for k, v in deal_record.items() 
                           if k in ['issuer_name', 'deal_short_name', 'as_of_date',
                                   'delinquency_30_plus', 'delinquency_60_plus', 'delinquency_90_plus',
                                   'extensions_rate', 'cnl_rate', 'oc_current', 'oc_target', 'oc_initial',
                                   'reserve_pct_original', 'reserve_pct_current']}
        
        if len(performance_data) > 3:
            self.performance_metrics_data.append(performance_data)
        
        # CNL Analysis Data
        cnl_data = {k: v for k, v in deal_record.items() 
                   if k.startswith('cnl_') or k in ['issuer_name', 'deal_short_name', 'as_of_date']}
        
        if len(cnl_data) > 3:
            self.cnl_analysis_data.append(cnl_data)
        
        # Trigger Analysis Data
        trigger_data = {k: v for k, v in deal_record.items() 
                       if k.startswith('trigger_') or k in ['issuer_name', 'deal_short_name', 'as_of_date']}
        
        if len(trigger_data) > 3:
            self.trigger_analysis_data.append(trigger_data)
    
    def _create_dataframes(self) -> Dict[str, pd.DataFrame]:
        """Convert collected data into pandas DataFrames."""
        dataframes = {}
        
        # Deal Master DataFrame
        if self.deal_master_data:
            dataframes['deal_master'] = pd.DataFrame(self.deal_master_data)
            dataframes['deal_master'] = self._clean_dataframe(dataframes['deal_master'])
        
        # Collateral Characteristics DataFrame
        if self.collateral_characteristics_data:
            dataframes['collateral_characteristics'] = pd.DataFrame(self.collateral_characteristics_data)
            dataframes['collateral_characteristics'] = self._clean_dataframe(dataframes['collateral_characteristics'])
        
        # Performance Metrics DataFrame
        if self.performance_metrics_data:
            dataframes['performance_metrics'] = pd.DataFrame(self.performance_metrics_data)
            dataframes['performance_metrics'] = self._clean_dataframe(dataframes['performance_metrics'])
        
        # Historical Performance DataFrame
        if self.historical_performance_data:
            dataframes['historical_performance'] = pd.DataFrame(self.historical_performance_data)
            dataframes['historical_performance'] = self._clean_dataframe(dataframes['historical_performance'])
        
        # CNL Analysis DataFrame
        if self.cnl_analysis_data:
            dataframes['cnl_analysis'] = pd.DataFrame(self.cnl_analysis_data)
            dataframes['cnl_analysis'] = self._clean_dataframe(dataframes['cnl_analysis'])
        
        # Trigger Analysis DataFrame
        if self.trigger_analysis_data:
            dataframes['trigger_analysis'] = pd.DataFrame(self.trigger_analysis_data)
            dataframes['trigger_analysis'] = self._clean_dataframe(dataframes['trigger_analysis'])
        
        # Statistical Summaries DataFrame
        if self.statistical_summaries_data:
            dataframes['statistical_summaries'] = pd.DataFrame(self.statistical_summaries_data)
            dataframes['statistical_summaries'] = self._clean_dataframe(dataframes['statistical_summaries'])
        
        return dataframes
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize a DataFrame."""
        # Remove duplicate rows
        df = df.drop_duplicates()
        
        # Convert date columns
        date_columns = ['as_of_date', 'securitization_date', 'prior_action_date', 'distribution_date']
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        
        # Convert numeric columns
        numeric_columns = [col for col in df.columns if any(keyword in col.lower() 
                          for keyword in ['balance', 'rate', 'factor', 'ratio', 'pct', 'cnl', 'delinquency', 'oc_'])]
        
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        return df

def main():
    """
    Main function to demonstrate usage of the ABS Data Extractor.
    """
    # Initialize the extractor
    extractor = ABSDataExtractor()
    
    # File paths (update these to your actual file paths)
    octane_file = "Octane Receivables Trust Comprehensive Surveillance Dashboard.xlsx"
    carvana_file = "Carvana Auto Receivables Trust Comprehensive Surveillance Dashboard.xlsx"
    
    try:
        # Extract data from both files
        dataframes = extractor.extract_from_files(octane_file, carvana_file)
        
        # Display summary of extracted data
        print("\n" + "="*50)
        print("ABS DATA EXTRACTION SUMMARY")
        print("="*50)
        
        for name, df in dataframes.items():
            print(f"\n{name.upper()}:")
            print(f"  Rows: {len(df)}")
            print(f"  Columns: {len(df.columns)}")
            print(f"  Columns: {list(df.columns)}")
            
            if len(df) > 0:
                print(f"  Sample data:")
                print(df.head(2).to_string())
        
        # Save to Excel files for review
        output_file = f"abs_extracted_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for name, df in dataframes.items():
                df.to_excel(writer, sheet_name=name[:31], index=False)  # Excel sheet name limit
        
        print(f"\nData saved to: {output_file}")
        
        return dataframes
        
    except Exception as e:
        logger.error(f"Error in main execution: {str(e)}")
        raise

if __name__ == "__main__":
    main()